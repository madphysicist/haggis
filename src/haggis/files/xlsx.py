#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# haggis: a library of general purpose utilities
#
# Copyright (C) 2019  Joseph R. Fox-Rabinovitz <jfoxrabinovitz at gmail dot com>
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

# Author: Joseph Fox-Rabinovitz <jfoxrabinovitz at gmail dot com>
# Version: 13 Apr 2019: Initial Coding


"""
Utilities for working with new-style MS Excel documents, only available
when the ``[xlsx]`` :ref:`extra <installation-extras>` is installed.

If `openpyxl`_ is not found at import time, this module will have a
:py:data:`xlsx_enabled` attribute, which will be :py:obj:`False`. If
`openpyxl`_ is found, on the other hand, :py:data:`xlsx_enabled` will
be :py:obj:`True`, and all the dependent functions and attributes of the
module will be present.

.. py:data:: xlsx_enabled

   A boolean value indicating whether the ``[xlsx]``
   :ref:`extra <installation-extras>` has been installed. If
   :py:obj:`False`, the API will be severely limited.


.. include:: /link-defs.rst 
"""

__all__ = ['xlsx_enabled', 'EXTENSION', 'ensure_extension',]


from warnings import warn

from . import ensure_extension as _ensure_extension


try:
    from openpyxl.styles import Border
except ImportError:
    from .. import _display_missing_extra
    _display_missing_extra('xlsx', 'openpyxl')
    xlsx_enabled = False
else:
    __all__.extend(['apply_border', 'copy_range'])
    xlsx_enabled = True


#: The extension that will be appended to output file names by
#: :py:func:`ensure_extension`.
EXTENSION = '.xlsx'


def ensure_extension(output):
    """
    Verify that the output object is a valid file name, and return a
    fixed version if not.

    If `output` is a :py:class:`str` and does not end with ``'.xlsx'``,
    fix it. If it ends with ``'.xls'``, append ``'x'``, otherwise append
    ``'.xlsx'``. All other types are assumed to be proper file-like
    objects that are passed through.
    """
    if isinstance(output, str):
        return _ensure_extension(output, EXTENSION, partial_policy='append',
                                 partial_limit=1)
    return output


if xlsx_enabled:
    def copy_range(ws_in, row_in=None, column_in=None, width=None, height=None,
                   ws_out=None, row_out=None, column_out=None, delete_empty=True):
        """
        Copy a range of cells from one worksheet to another.

        All arguments besides `ws_in` are technically optional, with
        some restrictions. This allows shortcuts for copying blocks
        between or even within worksheets.

        .. warning::

           At this time, copying into the same worksheet may yield
           incorrect results if the source and destination overlap.

        Parameters
        ----------
        ws_in : openpyxl.worksheet.worksheet.Worksheet
            The source worksheet with the data.
        row_in : int or None
            The row of the upper-left hand corner in the source. If
            :py:obj:`None`, single-column mode will be used. In this
            case, `width` may not be specified and efectively becomes
            ``1``.
        column_in : int or None
            The column of the upper-left hand corner of the source. If
            :py:obj:`None`, single-row mode will be used. In this case,
            ``height`` may not be specified and efectively becomes
            ``1``.
        width : int or None
             The width (number of columns) of the range. If
             :py:obj:`None`, all available columns will be used.
        height : int or None
            The height (number of rows) of the range. If :py:obj:`None`,
            all available rows will be used.
        ws_out : openpyxl.worksheet.worksheet.Worksheet
            The destination worksheet. If omitted, defaults to the
            source worksheet. In that case, the location being copied to
            must be different.
        row_out : int or None
            The row of the upper-left hand corner in the destination. If
            omitted, defaults to the same location as the source.
        column_out : int or None
            The column of the upper-left hand corner in the destination.
            If omitted, defaults to the same location as the source.
        delete_empty : bool
            Whether or not to remove destination cells that are
            overwritten by missing cells in the source range. Default is
            :py:obj:`True`.


        Only one of `row_in` and `column_in` may be :py:obj:`None`. All
        indices are one-based to conform to openpyxl notation.

        Return
        ------
        n : int
            The number of non-empty source cells copied. This will be zero
            if the destination is the same as the source.
        """
        if row_in is None and column_in is None:
            raise ValueError('Either `row_in` or `col_in` must be specified')

        if row_in is None:
            # Column-only mode
            if width is not None:
                raise ValueError('`width` not allowed for column copy')
            row_in = 1
            mr, mc = None, column_in
        elif column_in is None:
            # Row-only mode
            if height is not None:
                raise ValueError('`height` not allowed for row copy')
            column_in = 1
            mr, mc = row_in, None
        else:
            mr = None if height is None else row_in + height - 1
            mc = None if width is None else column_in + width - 1

        if row_out is None:
            row_out = row_in
        if column_out is None:
            column_out = column_in

        if ws_in == ws_out and row_in == row_out and column_in == column_out:
            warn('Source and destination are the same. Skipping.')
            return 0

        iterator = ws_in.iter_rows(min_row=row_in, max_row=mr,
                                   min_col=column_in, max_col=mc)

        counter = 0
        for dest_r, row in enumerate(iterator, start=row_out):
            for dest_c, item in enumerate(row, start=column_out):
                value = item.value
                copy = value is not None
                if copy or delete_empty:
                    ws_out.cell(row=dest_r, column=dest_c).value = value
                    if copy:
                        counter += 1

        return counter


    def apply_border(ws, start_row, end_row, start_column, end_column, *,
                     merge=False, **kwargs):
        """
        Apply a border around the specified range of cells as if it was
        a single object.

        The border may be specified in a number of ways through keyword
        arguments.

        If the argument `border` is specified, there may not be any
        other `kwargs`. If `border` is a
        :py:class:`openpyxl.styles.borders.Side`, apply it on all sides.
        If it is a :py:class:`openpyxl.styles.borders.Border`, apply the
        attributes of the border to the sides they correspond to around
        the whole edge.

        The other option is to specify `kwargs` as some subset of
        {`top`, `left`, `bottom`, `right`}. Each argument can be a
        :py:class:`openpyxl.styles.borders.Side` or an
        :py:class:`openpyxl.styles.borders.Border`. Only the part of the
        border corresponding to the side of the object it goes on will
        be used. For example, only the
        :py:attr:`~openpyxl.styles.borders.Border.top` attribute of a
        border specified for `top` will be used. The other attributes
        will be ignored. The borders corresponding to missing or
        :py:obj:`None` arguments will not be changed in this case.

        This function is based heavily on the recipe described in

        http://openpyxl.readthedocs.io/en/stable/styles.html#styling-merged-cells.
        """

        def check_side(side, direction):
            """
            If passed a `Side`, convert into a border with only that direction set.
            If passed a `Border`, effectively unset all the other directions.
            If `None`, pass-through.
            """
            if side is None:
                return side
            if isinstance(side, Border):
                side = getattr(side, direction)
            return Border(**{direction: side})

        if len(kwargs) == 1 and 'border' in kwargs:
            top = left = bottom = right = kwargs['border']
        elif set(kwargs) <= {'top', 'left', 'bottom', 'right'}:
            top = kwargs.get('top')
            left = kwargs.get('left')
            bottom = kwargs.get('bottom')
            right = kwargs.get('right')
        else:
            raise ValueError('Found illegal keyword arguments.')

        if merge:
            # At least in LibreOffice, the border does not come out right if it
            # is applied before the merge, so merge first, apply border after.
            ws.merge_cells(start_row=start_row, end_row=end_row,
                           start_column=start_column, end_column=end_column)

        top = check_side(top, 'top')
        left = check_side(left, 'left')
        bottom = check_side(bottom, 'bottom')
        right = check_side(right, 'right')

        for row in range(start_row, end_row + 1):
            if left:
                cell = ws.cell(row=row, column=start_column)
                cell.border = cell.border + left
            if right:
                cell = ws.cell(row=row, column=end_column)
                cell.border = cell.border + right

        for column in range(start_column, end_column + 1):
            if top:
                cell = ws.cell(row=start_row, column=column)
                cell.border = cell.border + top
            if bottom:
                cell = ws.cell(row=end_row, column=column)
                cell.border = cell.border + bottom
